#!/usr/bin/env python3
"""
Converts Kevin's VPH-_DATABASE.xlsx (the source-of-truth spreadsheet for The Valley Pocket
Harmonist, 2024) into indexes/VPH2024.json — the maintained, edition-specific page index
described in the Tunebook Registry spec (Section 10).

This does NOT touch tunebook-index.js. Capture and Compile only ever read `.title` per song,
and that's unaffected — this richer file is additive, for the Tunebook Editor and future
consumers, sitting alongside (not replacing) what's already in production.

Schema per song entry (keyed by page, using the suite's existing t/b split-page convention):
  title            — required, unchanged from today's tunebook-index.js shape
  firstLine
  meter            — kept as a free string; shape-note meter names have too many real
                     variants ("L. M.", "8, 7. D.", "C. M. D.") to usefully constrain
  timeSignature    — kept as a free string; a handful of tunes genuinely change meter
                     mid-piece ("4/4, 3/2") and that must not be silently split apart
  key
  textAttribution  — {credit, year}; "credit" is left as one opaque string rather than
                     split into named-author vs. source-book, because the spreadsheet
                     itself doesn't cleanly separate those (sometimes it's a person,
                     sometimes a book title, sometimes both across different verses)
  musicAttribution — {credit, year}; same reasoning — composer/arranger credit and
                     "alto by ..." credit are often one compound phrase, not separable
                     without inventing structure the source doesn't actually have
  source           — {citedAbbr, page}; omitted entirely when the spreadsheet's
                     "Source Abbr." was blank. This cites a HISTORICAL SOURCE WORK the
                     tune or text was drawn from (SZ, MCM, CHI, etc.) — a separate,
                     tune-level bibliography, not a Work/Edition Code from the tunebook
                     registry's own authority model. Deliberately named "citedAbbr" rather
                     than "workAbbr" so it can't be confused with the registry's own Work
                     Code terminology — they are unrelated code spaces. Not worth trying
                     to normalize or cross-reference against anything: each tunebook's
                     compiler invents these abbreviations independently, so there's no
                     consistent registry of them to check against, and per Kevin, building
                     one would be futile effort spent on something that doesn't matter to
                     what this suite actually does.

Normalization applied (visibly, not silently):
  - Page keys: "50T"/"50B" -> "50t"/"50b", matching the suite's existing convention.
  - Missing-value spellings ("Not specified", "Unknown", "N/A", None) all collapse to ""
    on both year fields, matching how tunebook-index.js already treats "" as "not known"
    for shmhaCode etc.
  - The spreadsheet's own "Order" column is dropped: it's just row position, fully
    redundant with the page key once sorted, and carrying it forward would just be a
    second, driftable copy of information the key already encodes.
"""

import json
import re
import sys
import openpyxl

SRC_XLSX = "/mnt/user-data/uploads/VPH-_DATABASE.xlsx"
OUT_JSON = "/home/claude/indexes/VPH2024.json"
EDITION_CODE = "VPH2024"
INDEX_VERSION = "1.0.0"

MISSING_SPELLINGS = {"not specified", "unknown", "n/a", "nan", ""}


def clean(v):
    """None / NaN / known missing-value spellings all become ''. Everything else: trimmed string."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in MISSING_SPELLINGS:
        return ""
    return s


def canon_page_key(raw_call):
    """'50T' -> '50t', '386B' -> '386b', '1' -> '1' — matches the suite's existing convention."""
    s = str(raw_call).strip()
    m = re.match(r"^(\d+)([TB])$", s)
    if m:
        return m.group(1) + m.group(2).lower()
    return s


def split_source(raw_abbr):
    """'CHI 50' -> {citedAbbr:'CHI', page:'50'}; 'SZ' -> {citedAbbr:'SZ', page:''}; None -> None."""
    if raw_abbr is None:
        return None
    s = str(raw_abbr).strip()
    if not s:
        return None
    m = re.match(r"^([A-Za-z0-9]+)(?:\s+(\d+))?$", s)
    if not m:
        # Shouldn't happen — every value in the source file matched this pattern when checked —
        # but fail loudly rather than silently mis-splitting if the data ever changes.
        raise ValueError("Source Abbr. value didn't match the expected pattern: %r" % raw_abbr)
    return {"citedAbbr": m.group(1), "page": m.group(2) or ""}


def main():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["Sheet1"]

    songs = {}
    row_count = 0
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        if not any(v is not None for v in vals):
            continue
        row_count += 1
        (call, title, meter, time_sig, key, source_abbr,
         text_author, text_year, composer, comp_year, first_line, _order) = vals

        page = canon_page_key(call)

        entry = {
            "title": clean(title),
            "firstLine": clean(first_line),
            "meter": clean(meter),
            "timeSignature": clean(time_sig),
            "key": clean(key),
            "textAttribution": {"credit": clean(text_author), "year": clean(text_year)},
            "musicAttribution": {"credit": clean(composer), "year": clean(comp_year)},
        }
        src = split_source(source_abbr)
        if src is not None:
            entry["source"] = src

        if page in songs:
            raise ValueError("Duplicate page key %r at row %d — refusing to silently overwrite." % (page, r))
        songs[page] = entry

    out = {
        "editionCode": EDITION_CODE,
        "indexVersion": INDEX_VERSION,
        "songs": songs,
    }

    import os
    os.makedirs("/home/claude/indexes", exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print("Converted %d rows -> %d song entries" % (row_count, len(songs)))
    print("Wrote %s" % OUT_JSON)


if __name__ == "__main__":
    main()
